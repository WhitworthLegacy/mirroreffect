#!/usr/bin/env python3
"""
Extract all sheets from the Excel file to CSV files
"""
import openpyxl
import csv
import os
from pathlib import Path

EXCEL_PATH = Path(__file__).parent.parent.parent.parent / "files" / "CLIENTS (1).xlsx"
OUTPUT_DIR = Path(__file__).parent.parent.parent.parent / "files" / "csv"

def extract_sheet_to_csv(workbook, sheet_name, output_path):
    """Extract a single sheet to CSV"""
    sheet = workbook[sheet_name]

    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)

        for row in sheet.iter_rows(values_only=True):
            # Skip completely empty rows
            if all(cell is None or cell == '' for cell in row):
                continue
            writer.writerow(row)

    print(f"✓ Extracted {sheet_name} to {output_path.name}")

def main():
    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Load workbook
    print(f"Loading workbook from {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # Extract each sheet
    for sheet_name in wb.sheetnames:
        output_path = OUTPUT_DIR / f"{sheet_name.lower()}.csv"
        extract_sheet_to_csv(wb, sheet_name, output_path)

    print(f"\n✓ All sheets extracted to {OUTPUT_DIR}")
    print("\nExtracted files:")
    for csv_file in sorted(OUTPUT_DIR.glob("*.csv")):
        size = csv_file.stat().st_size
        print(f"  - {csv_file.name} ({size:,} bytes)")

if __name__ == "__main__":
    main()
